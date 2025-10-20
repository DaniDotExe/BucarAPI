#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para agregar unidades a las columnas del archivo Excel de datos meteorológicos
Fecha: 2025-10-20
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os

def agregar_unidades_excel(archivo_entrada='datos_meteorologicos_bucaramanga.xlsx', 
                          archivo_salida='datos_meteorologicos_bucaramanga_con_unidades.xlsx'):
    """
    Lee el archivo Excel, agrega unidades a los nombres de las columnas y guarda el resultado
    """
    
    if not os.path.exists(archivo_entrada):
        print(f"Error: No se encontró el archivo '{archivo_entrada}'")
        print("Por favor, ejecuta primero Main.py para generar los datos.")
        return False
    
    print("=" * 70)
    print("AGREGANDO UNIDADES A LOS DATOS METEOROLÓGICOS")
    print("=" * 70)
    print(f"\nArchivo de entrada: {archivo_entrada}")
    
    try:
        # Leer el archivo Excel
        df = pd.read_excel(archivo_entrada)
        print(f"Registros leídos: {len(df)}")
        print(f"Columnas originales: {list(df.columns)}")
        
        # Diccionario con las unidades para cada columna
        unidades = {
            'Ciudad': '',  # Sin unidad
            'Fecha': '',   # Sin unidad
            'Hora': '',    # Sin unidad
            'Temperatura': '(°C)',
            'Presión': '(hPa)',
            'Humedad': '(%)',
            'Punto de Rocío': '(°C)',
            'Precipitación': '(mm)',
            'Nieve': '(mm)',
            'Dirección Viento': '(°)',
            'Velocidad Viento': '(km/h)',
            'Ráfaga Viento': '(km/h)',
            'Horas Sol': '(min)',
            'Condición': ''  # Sin unidad (es un código)
        }
        
        # Crear un diccionario para renombrar las columnas
        nuevos_nombres = {}
        for col in df.columns:
            if col in unidades:
                unidad = unidades[col]
                if unidad:
                    nuevos_nombres[col] = f"{col} {unidad}"
                else:
                    nuevos_nombres[col] = col
            else:
                # Si hay alguna columna que no está en el diccionario, la dejamos sin cambios
                nuevos_nombres[col] = col
        
        # Renombrar las columnas
        df.rename(columns=nuevos_nombres, inplace=True)
        
        print(f"\nColumnas con unidades: {list(df.columns)}")
        
        # Guardar el nuevo archivo Excel con formato
        with pd.ExcelWriter(archivo_salida, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Datos Meteorológicos')
            
            # Obtener el worksheet para aplicar formato
            worksheet = writer.sheets['Datos Meteorológicos']
            
            # Aplicar formato al encabezado
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Ajustar ancho de columnas
            for idx, col in enumerate(df.columns, 1):
                column_letter = worksheet.cell(row=1, column=idx).column_letter
                
                # Calcular el ancho basado en el contenido
                max_length = max(
                    df.iloc[:, idx-1].astype(str).apply(len).max() if len(df) > 0 else 0,
                    len(str(col))
                ) + 2
                
                worksheet.column_dimensions[column_letter].width = min(max_length, 25)
            
            # Congelar la primera fila
            worksheet.freeze_panes = 'A2'
        
        print(f"\n✓ Archivo guardado exitosamente: {archivo_salida}")
        print("=" * 70)
        
        # Mostrar resumen de cambios
        print("\nRESUMEN DE CAMBIOS:")
        print("-" * 70)
        cambios_realizados = []
        for col_original, col_nueva in nuevos_nombres.items():
            if col_original != col_nueva:
                cambios_realizados.append(f"  • {col_original:20} → {col_nueva}")
        
        if cambios_realizados:
            for cambio in cambios_realizados:
                print(cambio)
        else:
            print("  No se realizaron cambios en los nombres de las columnas")
        
        print("-" * 70)
        print(f"\nTotal de columnas procesadas: {len(df.columns)}")
        print(f"Total de registros: {len(df)}")
        print("\n✓ Proceso completado exitosamente")
        
        return True
        
    except Exception as e:
        print(f"\n✗ Error al procesar el archivo: {e}")
        import traceback
        traceback.print_exc()
        return False

def main():
    """
    Función principal
    """
    # Archivo de entrada (generado por Main.py)
    archivo_entrada = 'datos_meteorologicos_bucaramanga.xlsx'
    
    # Archivo de salida con unidades
    archivo_salida = 'datos_meteorologicos_bucaramanga_con_unidades.xlsx'
    
    # Procesar el archivo
    exito = agregar_unidades_excel(archivo_entrada, archivo_salida)
    
    if exito:
        print(f"\n{'='*70}")
        print("El archivo con unidades está listo para usar:")
        print(f"📊 {archivo_salida}")
        print(f"{'='*70}")
    else:
        print("\n❌ El proceso no se completó correctamente.")

if __name__ == "__main__":
    main()
